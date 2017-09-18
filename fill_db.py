from time import sleep
from datetime import datetime
from os.path import join
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook

from xlsx_read import beep
from geolocate import  COLS_XLSX#, MATCHED_XLSX

MATCHED_XLSX = join('data', 'rest.xlsx')

SERVER_URL = '*********' # remote server url
LOGIN = '********'
PASSWORD = '*********'

REPORT_TXT = join('data', 'report.txt')

USER_NAME_ID = 'userNameInput'
PASSWORD_ID = 'passwordInput'
LOGIN_SUBMIT_ID = 'submitButton'
OPEN_POS_ID = 'openStores'
SEARCH_ICON_CSS = 'span[data-bind="click:loadSearchFieldsList"]'
ADDRESS_ID_FIELD_CSS = 'input[field-name="DPA"]'
SEARCH_SUBMIT_CSS = 'div[data-bind="click:searchUserStores"]'
SALE_POINT_CSS = 'span[class="ui-gridBtn ui-subactions"]'
DSPV_LI_CSS = 'li[data-bind="click:$root.selectOrUnselectUser"]'
DSPV_NAME_CSS = 'span[class="ui-redirect-user"]'
RADIO_CSS = 'input[type="radio"]'
DSPV_SUBMIT_ID = 'btnRedirect'
DSPV_CANCEL_ID = 'btnCancelRedirect'

TIME_OUT_ALARM = 'Application Close'

def read_mappings():
    # create list of maps as dicts    
    col_id = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[COLS_XLSX.index('addressid')]
    col_dspv = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[COLS_XLSX.index('near_dspv')]
    wb = load_workbook(filename=MATCHED_XLSX)
    ws = wb.active
    map_list = []
    for i in range(1, ws.max_row + 1):
        map_list.append({'id': ws[col_id + str(i)].value, 'dspv': ws[col_dspv + str(i)].value})
    wb.close()
    return map_list


def fill_web_db(map_list):
    
    def login():
        # login at server
        driver = webdriver.Chrome()
        driver.get(SERVER_URL)
        try:
            driver.find_element_by_id(USER_NAME_ID).send_keys(LOGIN)
            driver.find_element_by_id(PASSWORD_ID).send_keys(PASSWORD)
            driver.find_element_by_id(LOGIN_SUBMIT_ID).click()
        except:
            pass
        while True:
            try:
                driver.find_element_by_id(OPEN_POS_ID).click()
            except:
                sleep(1)
            else:
                break
        return driver

    def transmit(report_file, message):
        # send message to report_file
        report_file.write(message + '\n')

    with open(REPORT_TXT, 'at', encoding='windows-1251')as f_rep:
        transmit(f_rep, 'Started: ' + str(datetime.now()))
        driver = login()
        # fill db
        for i, np in enumerate(map_list[1:]):
            while True:
                try:
                    transmit(f_rep, '{:>8})Matching point id#{} with DSPV {}: '.format(i + 1, *np.values()))
                    # SEARCH icon
                    while True:
                        try:
                            driver.find_element_by_css_selector(SEARCH_ICON_CSS).click()
                        except:
                            assert TIME_OUT_ALARM not in driver.title
                            sleep(1)
                        else:
                            break
                    # ADDRESS_ID field
                    while True:
                        try:
                            elem = driver.find_element_by_css_selector(ADDRESS_ID_FIELD_CSS)
                        except:
                            assert TIME_OUT_ALARM not in driver.title
                            sleep(1)
                        else:
                            break
                    while True:
                        try:
                            elem.clear()
                            elem.send_keys(np['id'])
                        except:
                            assert TIME_OUT_ALARM not in driver.title
                            sleep(1)
                        else:
                            break
                    # SEARCH submit button
                    while True:
                        try:
                            driver.find_element_by_css_selector(SEARCH_SUBMIT_CSS).click()
                        except:
                            assert TIME_OUT_ALARM not in driver.title
                            sleep(1)
                        else:
                            break
                    # select found sale point
                    sleep(10)
                    try:
                        driver.find_element_by_css_selector(SALE_POINT_CSS).click()
                    except:
                        assert TIME_OUT_ALARM not in driver.title
                        transmit(f_rep, 'this sale point does not exist')
                        break
                    # REDIRECT button
                    try:
                        driver.find_element_by_link_text('Redirect').click()
                    except NoSuchElementException:
                        assert TIME_OUT_ALARM not in driver.title
                        transmit(f_rep, 'there is not REDIRECT option')
                        break
                    # look for respective dspv in list and check it
                    while True:
                        try:
                            for li in driver.find_elements_by_css_selector(DSPV_LI_CSS):
                                if ((li.find_element_by_css_selector(DSPV_NAME_CSS)).text == np['dspv']):
                                    radio = li.find_element_by_css_selector(RADIO_CSS)
                                    while not radio.is_selected():
                                        sleep(1)
                                        radio.click()
                                    message = 'successfully accomplished'
                                    # REDIRECT SUBMIT button
                                    driver.find_element_by_id(DSPV_SUBMIT_ID).click()
                                    break
                            else:
                                message = 'DSPV {} does not exist'.format(np['dspv'])
                                # REDIRECT CANCEL button
                                driver.find_element_by_id(DSPV_CANCEL_ID).click()
                        except:
                            assert TIME_OUT_ALARM not in driver.title
                            sleep(5)
                        else:
                            break
                    transmit(f_rep, message)
                except AssertionError:
                    driver.close()
                    driver = login()
                else:
                    break
        driver.close()
        transmit(f_rep, 'Finished: ' + str(datetime.now()))


def main():
    map_list = read_mappings()
    beep(1)
    fill_web_db(map_list)
    beep(2)


if __name__ == '__main__':
    main()
