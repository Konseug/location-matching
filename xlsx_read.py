import pickle
from os.path import join
from openpyxl import load_workbook

CHIEFS_XLSX = join('data', 'GT_Pos-list.xlsx')
POINTS_XLSX = join('data', 'GT_Astra_Matching.xlsx')
PICKLE_PATH = join('data', 'pickle.dat')

CHIEFS_COLUMNS = {
	'dspv2': 'i',
	'custcode': 'a',
	'te': 'm',
	'city': 'r',
	'address': 'p',
	'custname': 'n',
	'custregname': 'o',
	'coord': 'al'}
POINTS_COLUMNS = {
	'posname': 'c',
	'address': 'd',
	'city': 'e',
	'clientid': 'f',
	'addressid': 'g'}


def get_ch_dict():
	# create list of existing points as dicts basing on excel file
	wb = load_workbook(filename=CHIEFS_XLSX)
	ws = wb.active
	old_list = []
	for i in range(2, ws.max_row + 1):
		old_point = {}
		for key in CHIEFS_COLUMNS.keys():
			old_point[key] = ws[CHIEFS_COLUMNS[key] + str(i)].value
		try:
			old_point['lon'] = float(old_point['coord'].split()[0])
			old_point['lat'] = float(old_point['coord'].split()[1])
		except ValueError:
			old_point['detect'] = False
		else:
			old_point['detect'] = True
		old_list.append(old_point)
	wb.close()
	return old_list


def get_np_dict():
	# create list of new points as dicts basing on excel file
	wb = load_workbook(filename=POINTS_XLSX)
	ws = wb.active
	new_list = []
	for i in range(6, ws.max_row + 1):
		new_point = {}
		for key in POINTS_COLUMNS.keys():
			new_point[key] = ws[POINTS_COLUMNS[key] + str(i)].value
		new_list.append(new_point)
	wb.close()
	return new_list


def xlsx_to_pickle(*args):
	# read data from XLSX-files
	with open(PICKLE_PATH, 'wb') as f:
		[pickle.dump(arg, f) for arg in args]


def beep(n):
	print('\a' * n)


def main():
	xlsx_to_pickle(get_ch_dict(), get_np_dict())
	beep(1)


if __name__ == '__main__':
	main()